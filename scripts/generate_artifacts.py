from pathlib import Path
from typing import Iterable

from markdown import markdown
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, Preformatted, SimpleDocTemplate, Spacer


ROOT = Path(__file__).resolve().parents[1]
SPEC_DIR = ROOT / "spec"
DOCS_DIR = ROOT / "docs"

DIMENSION_WEIGHTS = [
    ("EstratĂ©gia e GovernanĂ§a de IA", 15),
    ("Dados e Infraestrutura", 20),
    ("Talento e Cultura", 10),
    ("Desenvolvimento e OperaĂ§ĂŁo de IA (DevOps/MLOps)", 25),
    ("Ă‰tica, TransparĂŞncia e GestĂŁo de Risco", 20),
    ("Impacto Social e Valor", 10),
]

INDICATORS = [
    ("EstratĂ©gia e GovernanĂ§a de IA", "EG1", "Cobertura de polĂ­tica e estratĂ©gia de IA", "percentual", "maior_melhor", 30, 30, 50, 75, 90, False),
    ("EstratĂ©gia e GovernanĂ§a de IA", "EG2", "GovernanĂ§a do portfĂłlio de casos de uso de IA", "percentual", "maior_melhor", 30, 30, 50, 75, 90, False),
    ("EstratĂ©gia e GovernanĂ§a de IA", "EG3", "Rastreabilidade de decisĂµes e artefatos de IA", "percentual", "maior_melhor", 40, 30, 50, 75, 90, False),
    ("Dados e Infraestrutura", "DI1", "Cobertura de catĂˇlogo e linhagem de dados e modelos", "percentual", "maior_melhor", 35, 35, 55, 75, 90, False),
    ("Dados e Infraestrutura", "DI2", "Conformidade de proteĂ§ĂŁo de dados e acessos", "percentual", "maior_melhor", 40, 40, 60, 80, 95, True),
    ("Dados e Infraestrutura", "DI3", "Disponibilidade da infraestrutura crĂ­tica de IA", "percentual", "maior_melhor", 25, 97.0, 98.0, 99.0, 99.7, False),
    ("Talento e Cultura", "TC1", "Cobertura de papĂ©is crĂ­ticos de IA formalmente atribuĂ­dos", "percentual", "maior_melhor", 35, 30, 50, 75, 90, False),
    ("Talento e Cultura", "TC2", "CapacitaĂ§ĂŁo aplicada em governanĂ§a e operaĂ§ĂŁo de IA", "horas", "maior_melhor", 25, 6, 12, 20, 32, False),
    ("Talento e Cultura", "TC3", "AderĂŞncia Ă  segregaĂ§ĂŁo entre ideaĂ§ĂŁo e produĂ§ĂŁo", "percentual", "maior_melhor", 40, 30, 50, 75, 90, False),
    ("Desenvolvimento e OperaĂ§ĂŁo de IA (DevOps/MLOps)", "DO1", "Cobertura de especificaĂ§ĂŁo antes de cĂłdigo ou ajuste de modelo", "percentual", "maior_melhor", 25, 35, 55, 80, 95, False),
    ("Desenvolvimento e OperaĂ§ĂŁo de IA (DevOps/MLOps)", "DO2", "Cobertura de testes e validaĂ§Ăµes de IA na esteira", "percentual", "maior_melhor", 30, 35, 55, 80, 95, True),
    ("Desenvolvimento e OperaĂ§ĂŁo de IA (DevOps/MLOps)", "DO3", "Cobertura de observabilidade obrigatĂłria", "percentual", "maior_melhor", 25, 35, 55, 80, 95, True),
    ("Desenvolvimento e OperaĂ§ĂŁo de IA (DevOps/MLOps)", "DO4", "Tempo mĂ©dio de resposta a incidentes de IA (MTTR)", "horas", "menor_melhor", 20, 72, 36, 12, 4, False),
    ("Ă‰tica, TransparĂŞncia e GestĂŁo de Risco", "ER1", "Cobertura de avaliaĂ§ĂŁo Ă©tica e de risco algorĂ­tmico", "percentual", "maior_melhor", 40, 35, 60, 80, 95, True),
    ("Ă‰tica, TransparĂŞncia e GestĂŁo de Risco", "ER2", "TransparĂŞncia ao usuĂˇrio e ao afetado pela IA", "percentual", "maior_melhor", 25, 35, 55, 80, 95, False),
    ("Ă‰tica, TransparĂŞncia e GestĂŁo de Risco", "ER3", "Conformidade Ă©tica e regulatĂłria apĂłs auditoria", "percentual", "maior_melhor", 35, 60, 75, 90, 98, True),
    ("Impacto Social e Valor", "IV1", "Taxa de metas de valor atingidas por iniciativas de IA", "percentual", "maior_melhor", 40, 30, 50, 75, 90, False),
    ("Impacto Social e Valor", "IV2", "Cobertura de acessibilidade e inclusĂŁo nos serviĂ§os com IA", "percentual", "maior_melhor", 25, 35, 55, 80, 95, False),
    ("Impacto Social e Valor", "IV3", "ConfianĂ§a e satisfaĂ§ĂŁo do usuĂˇrio com serviĂ§os apoiados por IA", "escala_1_5", "maior_melhor", 35, 2.8, 3.4, 4.0, 4.5, False),
]


def style_sheet_header(sheet, row: int, fill_color: str = "1F4E78") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)


def build_workbook() -> None:
    wb = Workbook()
    ws_instructions = wb.active
    ws_instructions.title = "InstruĂ§Ăµes"
    ws_dimensions = wb.create_sheet("ParĂ˘metros_DimensĂµes")
    ws_indicators = wb.create_sheet("ParĂ˘metros_Indicadores")
    ws_eval = wb.create_sheet("AvaliaĂ§ĂŁo")
    ws_summary = wb.create_sheet("Resumo")

    ws_instructions["A1"] = "FACIN_IA - Planilha de AvaliaĂ§ĂŁo AutomĂˇtica"
    ws_instructions["A1"].font = Font(bold=True, size=14)
    instruction_lines = [
        "1. Preencha apenas as colunas 'Valor apurado', 'EvidĂŞncia principal' e 'ObservaĂ§Ăµes' na aba AvaliaĂ§ĂŁo.",
        "2. A nota automĂˇtica do indicador Ă© calculada a partir dos limiares parametrizados.",
        "3. A pontuaĂ§ĂŁo da dimensĂŁo Ă© ponderada pelos pesos internos dos indicadores.",
        "4. O Ă­ndice geral Ă© ponderado pelos pesos das seis dimensĂµes.",
        "5. A classificaĂ§ĂŁo final aplica as barreiras institucionais da PRODEMGE.",
    ]
    for idx, line in enumerate(instruction_lines, start=3):
        ws_instructions[f"A{idx}"] = line
    ws_instructions.column_dimensions["A"].width = 120

    ws_dimensions.append(["DimensĂŁo", "Peso_dimensĂŁo"])
    for row in DIMENSION_WEIGHTS:
        ws_dimensions.append(row)
    style_sheet_header(ws_dimensions, 1)
    autosize(ws_dimensions)

    ws_indicators.append([
        "DimensĂŁo",
        "CĂłdigo",
        "Indicador",
        "Unidade",
        "Polaridade",
        "Peso_indicador",
        "Limiar_N2",
        "Limiar_N3",
        "Limiar_N4",
        "Limiar_N5",
        "Indicador_crĂ­tico",
    ])
    for row in INDICATORS:
        ws_indicators.append(row)
    style_sheet_header(ws_indicators, 1)
    autosize(ws_indicators)

    ws_eval.append([
        "DimensĂŁo",
        "CĂłdigo",
        "Indicador",
        "Unidade",
        "Polaridade",
        "Peso_indicador",
        "Limiar_N2",
        "Limiar_N3",
        "Limiar_N4",
        "Limiar_N5",
        "Valor_apurado",
        "Nota_automĂˇtica",
        "Indicador_crĂ­tico",
        "EvidĂŞncia_principal",
        "ObservaĂ§Ăµes",
    ])
    for row_idx, row in enumerate(INDICATORS, start=2):
        dim, code, name, unit, polarity, weight, t2, t3, t4, t5, critical = row
        ws_eval.append([dim, code, name, unit, polarity, weight, t2, t3, t4, t5, None, None, critical, "", ""])
        if polarity == "maior_melhor":
            formula = (
                f'=IF(K{row_idx}="","",IF(K{row_idx}<G{row_idx},1,'
                f'IF(K{row_idx}<H{row_idx},2,IF(K{row_idx}<I{row_idx},3,'
                f'IF(K{row_idx}<J{row_idx},4,5)))))'
            )
        else:
            formula = (
                f'=IF(K{row_idx}="","",IF(K{row_idx}<=J{row_idx},5,'
                f'IF(K{row_idx}<=I{row_idx},4,IF(K{row_idx}<=H{row_idx},3,'
                f'IF(K{row_idx}<=G{row_idx},2,1)))))'
            )
        ws_eval[f"L{row_idx}"] = formula
    style_sheet_header(ws_eval, 1)
    autosize(ws_eval)

    ws_summary.append(["DimensĂŁo", "Peso_dimensĂŁo", "PontuaĂ§ĂŁo_dimensĂŁo"])
    for idx, (dimension, weight) in enumerate(DIMENSION_WEIGHTS, start=2):
        ws_summary[f"A{idx}"] = dimension
        ws_summary[f"B{idx}"] = weight
        ws_summary[f"C{idx}"] = (
            f'=IFERROR(SUMPRODUCT((\'AvaliaĂ§ĂŁo\'!$A$2:$A$20=A{idx})*\'AvaliaĂ§ĂŁo\'!$F$2:$F$20*\'AvaliaĂ§ĂŁo\'!$L$2:$L$20)'
            f'/SUMIF(\'AvaliaĂ§ĂŁo\'!$A$2:$A$20,A{idx},\'AvaliaĂ§ĂŁo\'!$F$2:$F$20),"")'
        )

    ws_summary["A10"] = "ĂŤndice geral"
    ws_summary["C10"] = "=SUMPRODUCT(B2:B7,C2:C7)/100"
    ws_summary["A11"] = "ClassificaĂ§ĂŁo base"
    ws_summary["C11"] = (
        '=IF(C10="","",IF(C10<=1.9,"NĂ­vel 1 - Inicial",'
        'IF(C10<=2.7,"NĂ­vel 2 - Em Desenvolvimento",'
        'IF(C10<=3.5,"NĂ­vel 3 - Otimizado",'
        'IF(C10<=4.3,"NĂ­vel 4 - Consolidado","NĂ­vel 5 - Estabelecido")))))'
    )
    ws_summary["A12"] = "Barreira nĂ­veis crĂ­ticos"
    ws_summary["C12"] = '=MIN(C3,C5,C6)'
    ws_summary["A13"] = "Menor pontuaĂ§ĂŁo geral"
    ws_summary["C13"] = '=MIN(C2:C7)'
    ws_summary["A14"] = "Menor nota de indicador crĂ­tico"
    ws_summary["C14"] = '=MINIFS(\'AvaliaĂ§ĂŁo\'!L2:L20,\'AvaliaĂ§ĂŁo\'!M2:M20,TRUE)'
    ws_summary["A15"] = "ClassificaĂ§ĂŁo final"
    ws_summary["C15"] = (
        '=IF(C10="","",IF(AND(C10>4.3,C13>=4,C14>=3),"NĂ­vel 5 - Estabelecido",'
        'IF(AND(C10>3.5,C12>=3.5),"NĂ­vel 4 - Consolidado",'
        'IF(C10>2.7,"NĂ­vel 3 - Otimizado",'
        'IF(C10>1.9,"NĂ­vel 2 - Em Desenvolvimento","NĂ­vel 1 - Inicial")))))'
    )
    style_sheet_header(ws_summary, 1)
    autosize(ws_summary)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

    workbook_path = DOCS_DIR / "FACIN_IA_Avaliacao_Automatica.xlsx"
    wb.save(workbook_path)


def markdown_to_html(title: str, md_text: str) -> str:
    body = markdown(md_text, extensions=["tables", "fenced_code", "toc"])
    css = """
    body { font-family: 'Segoe UI', Arial, sans-serif; margin: 36px auto; max-width: 980px; color: #16324f; }
    h1, h2, h3 { color: #0b3a66; }
    table { width: 100%; border-collapse: collapse; margin: 16px 0; font-size: 13px; }
    th, td { border: 1px solid #c8d5e3; padding: 8px; vertical-align: top; }
    th { background: #0b3a66; color: white; }
    code { background: #eef3f8; padding: 2px 4px; }
    blockquote { border-left: 4px solid #0b3a66; padding-left: 12px; color: #385269; }
    """
    return f"""<!doctype html>
<html lang=\"pt-BR\">
<head>
  <meta charset=\"utf-8\" />
  <title>{title}</title>
  <style>{css}</style>
</head>
<body>
{body}
</body>
</html>
"""


def extract_title(md_path: Path, md_text: str) -> str:
    for raw_line in md_text.splitlines():
        if raw_line.startswith("# "):
            return raw_line[2:].strip()
    return md_path.stem.replace("_", " ")


def markdown_to_pdf(md_text: str, pdf_path: Path) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        textColor="#0b3a66",
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor="#0b3a66",
        spaceAfter=8,
        spaceBefore=10,
    )
    body_style = ParagraphStyle(
        "CustomBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=13,
        spaceAfter=6,
    )
    mono_style = ParagraphStyle(
        "CustomMono",
        parent=styles["Code"],
        fontName="Courier",
        fontSize=8,
        leading=10,
        spaceAfter=6,
    )

    story = []
    first_heading_used = False
    for raw_line in md_text.splitlines():
        line = raw_line.rstrip()
        if not line:
            story.append(Spacer(1, 6))
            continue
        if line.startswith("# "):
            text = line[2:].strip()
            if not first_heading_used:
                story.append(Paragraph(text, title_style))
                first_heading_used = True
            else:
                story.append(Paragraph(text, heading_style))
            continue
        if line.startswith("## ") or line.startswith("### "):
            text = line.lstrip("#").strip()
            story.append(Paragraph(text, heading_style))
            continue
        if line.startswith("|") or line.startswith("$$"):
            story.append(Preformatted(line, mono_style))
            continue
        if line.startswith("-") or line[:2].isdigit():
            story.append(Paragraph(line.replace("&", "&amp;"), body_style))
            continue
        story.append(Paragraph(line.replace("&", "&amp;"), body_style))

    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    doc.build(story)


def convert_markdown_documents(paths: Iterable[Path]) -> None:
    for md_path in paths:
        md_text = md_path.read_text(encoding="utf-8")
        title = extract_title(md_path, md_text)
        md_output_path = DOCS_DIR / f"{md_path.stem}.md"
        html_text = markdown_to_html(title, md_text)
        html_path = DOCS_DIR / f"{md_path.stem}.html"
        pdf_path = DOCS_DIR / f"{md_path.stem}.pdf"
        md_output_path.write_text(md_text, encoding="utf-8")
        html_path.write_text(html_text, encoding="utf-8")
        markdown_to_pdf(md_text, pdf_path)


def main() -> None:
    DOCS_DIR.mkdir(exist_ok=True)
    build_workbook()
    convert_markdown_documents(sorted(SPEC_DIR.glob("*.md")))


if __name__ == "__main__":
    main()